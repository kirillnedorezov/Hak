import openpyxl
import csv

# Загрузка книги Excel
old_KSR = openpyxl.load_workbook("classification.xlsx")
sheet1 = old_KSR["Материалы, изд, констр и оборуд"]
sheet2 = old_KSR["Машины и механизмы"]

# Создание CSV файла для записи
with open("classification.csv", "w", encoding="utf-8") as new_KSR:
    file_writer = csv.writer(new_KSR, delimiter=",", lineterminator="\r")
    my_dict = {}

    # Обработка первого листа
    for row in sheet1.iter_rows(values_only=True):
        if None not in row:  # Пропуск строк с None (объединенные ячейки)
            # Добавление или обновление записи в словаре
            if row[1] not in my_dict or len(my_dict[row[1]][0]) < len(row[0]):
                my_dict[row[1]] = [row[0], row[2]]

    # Обработка второго листа
    for row in sheet2.iter_rows(values_only=True):
        if None not in row:  # Пропуск строк с None (объединенные ячейки)
            # Добавление или обновление записи в словаре
            if row[1] not in my_dict or len(my_dict[row[1]][0]) < len(row[0]):
                my_dict[row[1]] = [row[0], row[2]]

    # Запись словаря в CSV файл
    for key, value in my_dict.items():
        file_writer.writerow([value[0], key, value[1]])
